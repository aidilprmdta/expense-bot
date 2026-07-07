"""
handlers/export.py
────────────────────
Export data transaksi ke file Excel (.xlsx) atau CSV, dikirim
sebagai dokumen langsung di chat Telegram.

Penggunaan:
  /export                  → bulan ini, format .xlsx (default)
  /export csv              → bulan ini, format .csv
  /export hari             → hari ini
  /export minggu           → 7 hari terakhir
  /export bulan            → bulan ini
  /export tahun            → tahun ini
  /export semua            → semua data (tanpa filter tanggal)
  /export 06/2025          → bulan spesifik
  /export tahun csv        → kombinasi periode + format
"""

import io
import csv
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from telegram import Update
from telegram.ext import ContextTypes

from handlers.sheets import get_all_records
from handlers.rekap import _parse_tgl, resolve_periode, rupiah, EMOJI_KAT

logger = logging.getLogger(__name__)

HELP_TEXT = (
    "Pilihan:\n"
    "`/export` — bulan ini (.xlsx)\n"
    "`/export csv` — bulan ini (.csv)\n"
    "`/export hari` — hari ini\n"
    "`/export minggu` — 7 hari terakhir\n"
    "`/export bulan` — bulan ini\n"
    "`/export tahun` — tahun ini\n"
    "`/export semua` — semua data\n"
    "`/export 06/2025` — bulan spesifik\n"
    "`/export tahun csv` — kombinasi periode + format"
)

PERIODE_VALID = {"hari", "kemarin", "minggu", "bulan", "tahun", "semua"}
FORMAT_VALID  = {"xlsx", "excel", "csv"}

HEADER_FILL = PatternFill(start_color="21927A", end_color="21927A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
COLUMNS     = ["Tanggal", "Nama Item", "Kategori", "Harga", "Catatan"]


# ─────────────────────────────────────────────────────────────
# FILTER DATA
# ─────────────────────────────────────────────────────────────

def _filter_records(records: list[dict], periode: str, tgt_bulan=None, tgt_tahun=None):
    """
    Filter records sesuai periode. Return (filtered_records, label_periode).
    periode == "semua" → tidak difilter sama sekali.
    """
    if periode == "semua":
        return records, "Semua Data"

    start, end, label = resolve_periode(periode, tgt_bulan, tgt_tahun)

    hasil = []
    for r in records:
        tgl = _parse_tgl(str(r.get("Tanggal", "")))
        if tgl is not None and start <= tgl <= end:
            hasil.append(r)

    return hasil, label


# ─────────────────────────────────────────────────────────────
# BUILD FILE
# ─────────────────────────────────────────────────────────────

def _build_xlsx(records: list[dict], label: str) -> io.BytesIO:
    """Buat file .xlsx dengan sheet Transaksi + sheet Ringkasan per kategori."""
    wb = Workbook()

    # ── Sheet 1: Transaksi ────────────────────────────────────
    ws = wb.active
    ws.title = "Transaksi"
    ws.append(COLUMNS)
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    total = 0
    per_kategori: dict[str, int] = {}

    for r in records:
        harga = _safe_int(r.get("Harga", 0))
        kategori = str(r.get("Kategori", "Lainnya")).strip() or "Lainnya"
        ws.append([
            r.get("Tanggal", ""),
            r.get("Nama Item", ""),
            kategori,
            harga,
            r.get("Catatan", "-"),
        ])
        total += harga
        per_kategori[kategori.lower()] = per_kategori.get(kategori.lower(), 0) + harga

    # Lebar kolom rapi
    for i, width in enumerate([14, 28, 16, 16, 24], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Format kolom Harga sebagai angka dengan pemisah ribuan
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = "#,##0"

    ws.freeze_panes = "A2"

    # ── Sheet 2: Ringkasan per kategori ────────────────────────
    # Pemasukan dipisah dari pengeluaran — keduanya beda sifat,
    # jadi tidak dijumlah jadi satu "total" yang membingungkan.
    pemasukan_total = per_kategori.pop("pemasukan", 0)
    pengeluaran_total = sum(per_kategori.values())

    ws2 = wb.create_sheet("Ringkasan")
    ws2.append(["Kategori", "Total", "Persentase dari Pengeluaran"])
    for col in range(1, 4):
        cell = ws2.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for kat, jml in sorted(per_kategori.items(), key=lambda x: x[1], reverse=True):
        persen = jml / pengeluaran_total if pengeluaran_total > 0 else 0
        ws2.append([kat.title(), jml, persen])

    ws2.append([])
    ws2.append(["TOTAL PENGELUARAN", pengeluaran_total, 1.0 if pengeluaran_total > 0 else 0])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True)

    if pemasukan_total > 0:
        ws2.append(["TOTAL PEMASUKAN", pemasukan_total, None])
        ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, color="21927A")
        ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True, color="21927A")

        ws2.append(["SELISIH (Pemasukan - Pengeluaran)", pemasukan_total - pengeluaran_total, None])
        ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
        ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True)

    for row in range(2, ws2.max_row + 1):
        cell_total = ws2.cell(row=row, column=2)
        if isinstance(cell_total.value, (int, float)):
            cell_total.number_format = "#,##0"
        cell_pct = ws2.cell(row=row, column=3)
        if isinstance(cell_pct.value, (int, float)):
            cell_pct.number_format = "0.0%"

    for i, width in enumerate([28, 18, 22], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_csv(records: list[dict]) -> io.BytesIO:
    """Buat file .csv sederhana, satu baris per transaksi."""
    text_buf = io.StringIO()
    writer = csv.writer(text_buf)
    writer.writerow(COLUMNS)

    for r in records:
        writer.writerow([
            r.get("Tanggal", ""),
            r.get("Nama Item", ""),
            r.get("Kategori", ""),
            _safe_int(r.get("Harga", 0)),
            r.get("Catatan", "-"),
        ])

    buf = io.BytesIO(text_buf.getvalue().encode("utf-8-sig"))  # BOM: Excel baca UTF-8 dgn benar
    buf.seek(0)
    return buf


def _safe_int(v) -> int:
    """Parse nilai Harga ke int, tahan format Sheets Indonesia ('25.000')."""
    if isinstance(v, (int, float)):
        return max(0, int(v))
    try:
        cleaned = str(v).strip().replace("Rp", "").replace(" ", "")
        if "." in cleaned and "," not in cleaned:
            parts = cleaned.split(".")
            if all(len(p) <= 3 for p in parts[1:]):
                cleaned = cleaned.replace(".", "")
        cleaned = cleaned.replace(",", "")
        return max(0, int(float(cleaned)))
    except (ValueError, TypeError):
        return 0


# ─────────────────────────────────────────────────────────────
# TELEGRAM HANDLER
# ─────────────────────────────────────────────────────────────

async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handler Telegram untuk /export."""
    args = [a.lower().strip() for a in (context.args or [])]

    periode   = "bulan"
    fmt       = "xlsx"
    tgt_bulan = tgt_tahun = None

    for arg in args:
        if "/" in arg:
            try:
                parts     = arg.split("/")
                tgt_bulan = int(parts[0])
                tgt_tahun = int(parts[1])
                if not (1 <= tgt_bulan <= 12) or tgt_tahun < 2020:
                    raise ValueError
                periode = "bulan"
            except (ValueError, IndexError):
                await update.message.reply_text(
                    "⚠️ Format bulan tidak valid.\nContoh: `/export 06/2025`",
                    parse_mode="Markdown",
                )
                return
        elif arg in PERIODE_VALID:
            periode = arg
        elif arg in FORMAT_VALID:
            fmt = "csv" if arg == "csv" else "xlsx"
        else:
            await update.message.reply_text(
                f"⚠️ Argumen `{arg}` tidak dikenal.\n\n{HELP_TEXT}",
                parse_mode="Markdown",
            )
            return

    loading = await update.message.reply_text("📦 Menyiapkan file export... ⏳")

    try:
        records = await get_all_records()
    except Exception as e:
        logger.error(f"[/export] Gagal baca Sheets: {e}", exc_info=True)
        await loading.edit_text(
            f"😔 *Gagal membaca Google Sheets*\n\n`{type(e).__name__}: {str(e)[:120]}`",
            parse_mode="Markdown",
        )
        return

    try:
        filtered, label = _filter_records(records, periode, tgt_bulan, tgt_tahun)

        if not filtered:
            await loading.edit_text(
                f"📦 *Export {label}*\n\n_Tidak ada transaksi di periode ini._",
                parse_mode="Markdown",
            )
            return

        label_file = label.replace(" ", "_")
        timestamp  = datetime.now().strftime("%Y%m%d")

        if fmt == "csv":
            file_buf = _build_csv(filtered)
            filename = f"Expense_{label_file}_{timestamp}.csv"
        else:
            file_buf = _build_xlsx(filtered, label)
            filename = f"Expense_{label_file}_{timestamp}.xlsx"

        total = sum(_safe_int(r.get("Harga", 0)) for r in filtered)

        await update.message.reply_document(
            document=file_buf,
            filename=filename,
            caption=(
                f"📦 *Export {label}*\n"
                f"📄 {len(filtered)} transaksi\n"
                f"💰 Total: {rupiah(total)}"
            ),
            parse_mode="Markdown",
        )
        await loading.delete()

    except Exception as e:
        logger.error(f"[/export] Gagal generate file: {e}", exc_info=True)
        await loading.edit_text(
            f"😔 *Gagal membuat file export*\n\n`{type(e).__name__}: {str(e)[:120]}`",
            parse_mode="Markdown",
        )


# ─────────────────────────────────────────────────────────────
# TEST MANUAL: python -m handlers.export
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    DUMMY = [
        {"Tanggal": "01/07/2026", "Nama Item": "Kopi", "Kategori": "Makan", "Harga": 25000, "Catatan": "-"},
        {"Tanggal": "02/07/2026", "Nama Item": "Bensin", "Kategori": "Transport", "Harga": 80000, "Catatan": "-"},
        {"Tanggal": "03/07/2026", "Nama Item": "Gaji", "Kategori": "Pemasukan", "Harga": 5000000, "Catatan": "-"},
    ]

    xlsx_buf = _build_xlsx(DUMMY, "Test Juli 2026")
    with open("test_export.xlsx", "wb") as f:
        f.write(xlsx_buf.read())
    print("✅ test_export.xlsx dibuat")

    csv_buf = _build_csv(DUMMY)
    with open("test_export.csv", "wb") as f:
        f.write(csv_buf.read())
    print("✅ test_export.csv dibuat")
